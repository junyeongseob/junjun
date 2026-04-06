from flask import Flask, send_from_directory, jsonify, request
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# 🔥 DB 경로 안정화 (Render 대응)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "attendance.db")

# 🔥 DB 초기화 (테이블 없으면 자동 생성)
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # 근무표 테이블
    cur.execute("""
    CREATE TABLE IF NOT EXISTS work_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        date TEXT,
        status TEXT
    )
    """)

    # 비상근무 테이블
    cur.execute("""
    CREATE TABLE IF NOT EXISTS special_duty (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        duty TEXT,
        name TEXT,
        date TEXT
    )
    """)

    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# 🔥 서버 시작 시 DB 초기화 실행
init_db()

# 근무표 조회
@app.route("/schedule")
def get_schedule():
    dates = request.args.get("dates")
    month = request.args.get("month")

    conn = get_db()
    cur = conn.cursor()

    if dates:
        date_list = dates.split(",")
        placeholders = ",".join("?" * len(date_list))
        cur.execute(
            f"SELECT name, date, status FROM work_schedule WHERE date IN ({placeholders})",
            date_list
        )
    elif month:
        cur.execute(
            "SELECT name, date, status FROM work_schedule WHERE date LIKE ?",
            (f"{month}%",)
        )
    else:
        return jsonify({})

    rows = cur.fetchall()
    conn.close()

    result = {}
    for r in rows:
        result.setdefault(r["name"], {})[r["date"]] = r["status"]

    return jsonify(result)

# 비상근무 조회
@app.route("/special")
def get_special():
    month = request.args.get("month")

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT duty, name, date FROM special_duty WHERE date LIKE ?",
        (f"{month}%",)
    )

    rows = cur.fetchall()
    conn.close()

    result = {}
    for r in rows:
        result.setdefault(r["date"], {}).setdefault(r["duty"], []).append(r["name"])

    return jsonify(result)

# 근무 추가
@app.route("/add_schedule", methods=["POST"])
def add_schedule():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
        (data["name"], data["date"], data["status"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 근무 일괄 추가
@app.route("/add_schedule_bulk", methods=["POST"])
def add_schedule_bulk():
    data = request.json["data"]

    conn = get_db()
    cur = conn.cursor()

    for line in data:
        if "\t" in line:
            parts = line.split("\t")
        else:
            parts = line.split(",")

        if len(parts) != 3:
            continue

        name, date, status = [p.strip() for p in parts]

        if not name or not date or not status:
            continue

        cur.execute(
            "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
            (name, date, status)
        )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 비상근무 추가
@app.route("/add_special", methods=["POST"])
def add_special():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO special_duty (duty, name, date) VALUES (?, ?, ?)",
        (data["duty"], data["name"], data["date"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 근무 삭제
@app.route("/delete_schedule", methods=["POST"])
def delete_schedule():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM work_schedule WHERE name=? AND date=?",
        (data["name"], data["date"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 비상근무 삭제
@app.route("/delete_special", methods=["POST"])
def delete_special():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM special_duty WHERE name=? AND date=? AND duty=?",
        (data["name"], data["date"], data["duty"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 🔥 기존 근무표 데이터 전체 삭제 (한 번만 쓰면 됨)
@app.route("/clear_schedule", methods=["POST"])
def clear_schedule():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM work_schedule")

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

# 🔥 엑셀 자동 업로드
@app.route("/upload_excel_auto", methods=["POST"])
def upload_excel_auto():
    file = request.files.get("file")

    if not file:
        return jsonify({"message": "파일 없음"}), 400

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    conn = get_db()
    cur = conn.cursor()

    inserted = 0

    # 남산분소 양식 기준
    header_row = 4       # 근무지 제목 행
    start_data_row = 5   # 실제 데이터 시작 행
    date_col = 1         # 날짜 열
    work_cols = range(3, ws.max_column + 1)

    ignore_headers = {"요일", "일자", "주요순찰지", "비고", ""}

    for row in range(start_data_row, ws.max_row + 1):
        raw_date = ws.cell(row=row, column=date_col).value

        if not raw_date:
            continue

        # 🔥 날짜 형식 통일
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            text_date = str(raw_date).strip()

            if text_date in ["", "None", "nan"]:
                continue

            # 4/8 → 2026-04-08 형태로 보정
            if "/" in text_date and len(text_date.split("/")) == 2:
                try:
                    month, day = text_date.split("/")
                    date_str = f"2026-{int(month):02d}-{int(day):02d}"
                except:
                    date_str = text_date
            else:
                date_str = text_date

        for col in work_cols:
            header_value = ws.cell(row=header_row, column=col).value
            cell_value = ws.cell(row=row, column=col).value

            if not cell_value:
                continue

            workplace = str(header_value).strip() if header_value else f"col_{col}"

            if workplace in ignore_headers:
                continue

            text = str(cell_value).strip()

            if text in ["", "-", "None", "nan"]:
                continue

            # 🔥 셀 안 이름 분리
            names = []
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                for name in line.split():
                    name = name.strip()
                    if name:
                        # 🔥 이름 정리
                        name = name.replace(",", "").replace("·", "").replace("/", "")
                        if name not in ["-", "및", "nan", "None", "(", ")"]:
                            names.append(name)

            for name in names:
                cur.execute(
                    "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
                    (name, date_str, workplace)
                )
                inserted += 1

    conn.commit()
    conn.close()

    return jsonify({"message": f"{inserted}개 자동 입력 완료"})

# HTML
@app.route("/")
def index():
    return send_from_directory(os.getcwd(), "test.html")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
